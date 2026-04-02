import json
from datetime import datetime
from decimal import Decimal

from django.db import transaction
from openpyxl import load_workbook

from .models import (
    Batch,
    ChemicalRecord,
    ChemicalType,
    EnvironmentRecord,
    EquipmentStatus,
    ProcessRecord,
    ResultRecord,
    RFTFlag,
)


def _normalize(value):
    if value in ("", None):
        return None
    return value


def _parse_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(str(value), fmt)
        except ValueError:
            continue
    return None


def _parse_json(value):
    if not value:
        return None
    if isinstance(value, (dict, list)):
        return value
    return json.loads(value)


def _safe_choice(value, choices, default):
    valid_values = {choice[0] for choice in choices}
    return value if value in valid_values else default


def import_from_excel(file_obj):
    wb = load_workbook(file_obj, data_only=True)

    batch_sheet = wb["批次主表"]
    chemical_sheet = wb["化学品表"]
    process_sheet = wb["工艺时序表"]
    env_sheet = wb["环境设备表"]
    result_sheet = wb["结果表"]

    imported_batches = set()

    with transaction.atomic():
        # 1) 批次主表
        for row in batch_sheet.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            batch, _ = Batch.objects.update_or_create(
                batch_id=str(row[0]).strip(),
                defaults={
                    "order_no": _normalize(row[1]) or "",
                    "fabric_weight_gsm": row[2],
                    "width_cm": row[3],
                    "fiber_type": row[4],
                    "fiber_structure": _normalize(row[5]) or "",
                    "weave_type": row[6],
                    "pretreatment_score": _normalize(row[7]),
                    "liquor_ratio": row[8],
                    "load_amount_kg": row[9],
                    "operator_name": row[10],
                    "machine_id": row[11],
                    "created_at": _parse_datetime(row[12]),
                },
            )
            imported_batches.add(batch.batch_id)

        # 避免重复明细
        ChemicalRecord.objects.filter(batch_id__in=imported_batches).delete()
        ProcessRecord.objects.filter(batch_id__in=imported_batches).delete()

        # 2) 化学品表
        for row in chemical_sheet.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            batch = Batch.objects.get(pk=str(row[0]).strip())
            ChemicalRecord.objects.create(
                batch=batch,
                chemical_type=row[1],
                chemical_name=row[2],
                concentration_gpl=row[3],
                measurement_method=_normalize(row[4]) or "",
                remark=_normalize(row[5]) or "",
            )

        # 3) 工艺时序表
        for row in process_sheet.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            batch = Batch.objects.get(pk=str(row[0]).strip())
            ProcessRecord.objects.create(
                batch=batch,
                time_min=row[1],
                temperature_c=row[2],
                heating_rate_cpm=_normalize(row[3]),
                ph=_normalize(row[4]),
                conductivity_ms_cm=_normalize(row[5]),
                flow_rate_lpm=_normalize(row[6]),
                dye_concentration_gpl=_normalize(row[7]),
                dye_uptake_pct=_normalize(row[8]),
                stirring_intensity_rpm=_normalize(row[9]),
                dye_concentration_spectrum=_parse_json(row[10]) if row[10] else None,
                remark=_normalize(row[11]) or "",
            )

        # 4) 环境设备表
        for row in env_sheet.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            batch = Batch.objects.get(pk=str(row[0]).strip())
            EnvironmentRecord.objects.update_or_create(
                batch=batch,
                defaults={
                    "water_ph": _normalize(row[1]),
                    "water_conductivity_us_cm": _normalize(row[2]),
                    "water_hardness_mgL": _normalize(row[3]),
                    "water_cod_mgL": _normalize(row[4]),
                    "equipment_status": row[5],
                    "vibration_mm_s": _normalize(row[6]),
                    "equipment_temp_c": _normalize(row[7]),
                    "sop_compliance": _normalize(row[8]) or "",
                    "remark": _normalize(row[9]) or "",
                },
            )

        # 5) 结果表
        for row in result_sheet.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            batch = Batch.objects.get(pk=str(row[0]).strip())
            ResultRecord.objects.update_or_create(
                batch=batch,
                defaults={
                    "dye_time_min": row[1],
                    "ks_value": _normalize(row[2]),
                    "reflectance_pct": _normalize(row[3]),
                    "delta_e_2000": row[4],
                    "spectrum_curve_json": _parse_json(row[5]) if row[5] else None,
                    "result_l": _normalize(row[6]),
                    "result_a": _normalize(row[7]),
                    "result_b": _normalize(row[8]),
                    "rft_flag": row[9],
                    "rework_count": _normalize(row[10]) or 0,
                    "energy_steam_kg": _normalize(row[11]),
                    "remark": _normalize(row[12]) or "",
                },
            )

    return {
        "batch_count": len(imported_batches),
        "chemical_count": ChemicalRecord.objects.filter(batch_id__in=imported_batches).count(),
        "process_count": ProcessRecord.objects.filter(batch_id__in=imported_batches).count(),
    }


def import_from_json_payload(payload):
    batch_id = str(payload.get("batch_id", "")).strip()
    if not batch_id:
        raise ValueError("batch_id 不能为空")

    chemicals_payload = payload.get("chemicals") or []
    process_payload = payload.get("process_records") or []
    env_payload = payload.get("environment") or {}
    result_payload = payload.get("result") or {}

    with transaction.atomic():
        batch, _ = Batch.objects.update_or_create(
            batch_id=batch_id,
            defaults={
                "order_no": payload.get("order_no", "") or "",
                "fabric_weight_gsm": payload.get("fabric_weight_gsm") or Decimal("0"),
                "width_cm": payload.get("width_cm") or Decimal("0"),
                "fiber_type": payload.get("fiber_type") or "PET",
                "fiber_structure": payload.get("fiber_structure", "") or "",
                "weave_type": payload.get("weave_type") or "平纹",
                "pretreatment_score": payload.get("pretreatment_score"),
                "liquor_ratio": payload.get("liquor_ratio") or Decimal("0"),
                "load_amount_kg": payload.get("load_amount_kg") or Decimal("0"),
                "operator_name": payload.get("operator_name") or "",
                "machine_id": payload.get("machine_id") or "",
                "created_at": _parse_datetime(payload.get("created_at")),
            },
        )

        ChemicalRecord.objects.filter(batch=batch).delete()
        for item in chemicals_payload:
            ChemicalRecord.objects.create(
                batch=batch,
                chemical_type=_safe_choice(
                    item.get("chemical_type"),
                    ChemicalType.choices,
                    ChemicalType.DYE,
                ),
                chemical_name=item.get("chemical_name") or "未命名化学品",
                concentration_gpl=item.get("concentration_gpl") or Decimal("0"),
                measurement_method=item.get("measurement_method") or "",
                remark=item.get("remark") or "",
            )

        ProcessRecord.objects.filter(batch=batch).delete()
        for idx, item in enumerate(process_payload):
            ProcessRecord.objects.create(
                batch=batch,
                time_min=item.get("time_min") if item.get("time_min") is not None else idx,
                temperature_c=item.get("temperature_c") or Decimal("0"),
                heating_rate_cpm=_normalize(item.get("heating_rate_cpm")),
                ph=_normalize(item.get("ph")),
                conductivity_ms_cm=_normalize(item.get("conductivity_ms_cm")),
                flow_rate_lpm=_normalize(item.get("flow_rate_lpm")),
                dye_concentration_gpl=_normalize(item.get("dye_concentration_gpl")),
                dye_uptake_pct=_normalize(item.get("dye_uptake_pct")),
                stirring_intensity_rpm=_normalize(item.get("stirring_intensity_rpm")),
                dye_concentration_spectrum=item.get("dye_concentration_spectrum"),
                remark=item.get("remark") or "",
            )

        EnvironmentRecord.objects.update_or_create(
            batch=batch,
            defaults={
                "water_ph": _normalize(env_payload.get("water_ph")),
                "water_conductivity_us_cm": _normalize(env_payload.get("water_conductivity_us_cm")),
                "water_hardness_mgL": _normalize(env_payload.get("water_hardness_mgL")),
                "water_cod_mgL": _normalize(env_payload.get("water_cod_mgL")),
                "equipment_status": _safe_choice(
                    env_payload.get("equipment_status"),
                    EquipmentStatus.choices,
                    EquipmentStatus.NORMAL,
                ),
                "vibration_mm_s": _normalize(env_payload.get("vibration_mm_s")),
                "equipment_temp_c": _normalize(env_payload.get("equipment_temp_c")),
                "sop_compliance": env_payload.get("sop_compliance") or "",
                "remark": env_payload.get("remark") or "",
            },
        )

        ResultRecord.objects.update_or_create(
            batch=batch,
            defaults={
                "dye_time_min": result_payload.get("dye_time_min") or 0,
                "ks_value": _normalize(result_payload.get("ks_value")),
                "reflectance_pct": _normalize(result_payload.get("reflectance_pct")),
                "delta_e_2000": result_payload.get("delta_e_2000") or Decimal("0"),
                "spectrum_curve_json": result_payload.get("spectrum_curve_json"),
                "result_l": _normalize(result_payload.get("result_l")),
                "result_a": _normalize(result_payload.get("result_a")),
                "result_b": _normalize(result_payload.get("result_b")),
                "rft_flag": _safe_choice(result_payload.get("rft_flag"), RFTFlag.choices, RFTFlag.YES),
                "rework_count": result_payload.get("rework_count") or 0,
                "energy_steam_kg": _normalize(result_payload.get("energy_steam_kg")),
                "remark": result_payload.get("remark") or "",
            },
        )

    return {
        "batch_id": batch_id,
        "chemical_count": len(chemicals_payload),
        "process_count": len(process_payload),
    }
